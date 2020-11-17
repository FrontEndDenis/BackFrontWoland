import datetime
import re
from urllib.request import urlopen

from PIL import Image
from django.http import Http404
from django.shortcuts import render, get_object_or_404
from django.views.generic.base import TemplateView
from transliterate import translit

from menu.models import Marka
from menu.models import MenuCatalog, Product  # , tableProductDop, ProductUsluga
from menu.models import Standart
from project_settings.models import ProjectSettings
from .forms import UploadFileForm
from .models import ImportData  # , ExportData

DEFAULT_OPTION = u'Все'

IND_STATE_WORK = 1
IND_STATE_READY = 2
IND_STATE_ERROR = 3
IND_STATE_POTENTIAL_ERROR = 4

ID_TYPE_MENU_CATALOG = 2

COUNT_ROW_FOR_INFO_UPDATE = 100
COUNT_ROW_FOR_INFO_UPDATE_CONTROL = 100
# COUNT_ROW_FOR_INFO_UPDATE_FORMAT = 1000

MAX_COUNT_ROW_INFO = 1000

TYPE_ACTION_IMPORT = "import"
TYPE_ACTION_IMPORT_MARKA = "import_marka"
TYPE_ACTION_IMPORT_HARD = "import_hard"
TYPE_ACTION_CONTROL = "control"

DEFAULT_MESSAGE_ERROR = u"Импорт завершен с ошибками"

COLUMN_ID = 0
COLUMN_ORDER = 1
COLUMN_NAME = 2
COLUMN_SIZE_A = 3
COLUMN_SIZE_B = 4
COLUMN_SIZE_C = 5
COLUMN_SIZE_D = 6
COLUMN_SIZE_E = 7
COLUMN_SIZE_F = 8
COLUMN_SIZE_L = 9
COLUMN_MARKA = 10
COLUMN_GOST = 11
COLUMN_AVAILABLE = 12
COLUMN_PRICE = 13
COLUMN_ED_IZM = 14
COLUMN_IMAGE_1 = 15
COLUMN_IMAGE_2 = 16
COLUMN_DESCRIPTION = 17
COLUMN_VENDOR = 18
COLUMN_COUNTRY_VENDOR = 19
COLUMN_METHOD = 20
COLUMN_CATEGORY_ID = 21
COLUMN_SERTIFICATE_ID = 22
COLUMN_PRODUCT_1_ID = 23
COLUMN_PRODUCT_2_ID = 24
COLUMN_PRODUCT_3_ID = 25
COLUMN_PRODUCT_4_ID = 26
COLUMN_USLUGA_1_ID = 27
COLUMN_USLUGA_2_ID = 28
COLUMN_USLUGA_3_ID = 29
COLUMN_USLUGA_4_ID = 30

COLUMN_IMAGE_1_HTTP = -2
COLUMN_IMAGE_2_HTTP = -1
COLUMN_IMAGE_1_TECH_NAME = -4
COLUMN_IMAGE_2_TECH_NAME = -3

COLUMN_COUNT = 31
ROW_MIN_COUNT = 2

CONTROL_CODE = "73aoF6N"

FLAG_IMPORT = 1


def static_admin_url(request):
	return {
		'static_admin_url': '/static/admin_m/',
	}


def import_image(request, image_name, url_image, info=None, product_id=None):
	from admin_m.views import IND_STATE_ERROR
	path_image = "www/media/uploads/images/load/" + image_name
	out = open(path_image, 'wb')
	
	if url_image[:4] == 'http':
		try:
			resource = urlopen(url_image)
		except:
			if info:
				info.info += u"ОШИБКА при загрузке изображения продукта c id=%d (проверьте правильность пути до изображения: %s)\n" % (
					int(product_id), url_image)
				# info.state_id = IND_STATE_ERROR
				info.save()
				return
	else:
		try:
			resource = urlopen('https://' + url_image)
			if resource.code == 404:
				resource = urlopen('http://' + url_image)
		except:
			try:
				resource = urlopen('http://' + url_image)
			except:
				if info:
					info.info += u"ОШИБКА при загрузке изображения продукта c id=%d (изображение содержит ошибки) %s\n" % (
						int(product_id), url_image)
					info.state_id = IND_STATE_ERROR
					info.save()
				return
	
	if resource.code == 404:
		if info:
			info.info += u"ОШИБКА 404 при загрузке изображения продукта c id=%d (изображение не найдено) %s\n" % (
				int(product_id), url_image)
			info.state_id = IND_STATE_ERROR
			info.save()
		return
	
	out.write(resource.read())
	out.close()
	try:
		img = Image.open(path_image).convert('RGB')
	except:
		resource = urlopen('http://' + url_image)
		out = open(path_image, 'wb')
		out.write(resource.read())
		out.close()
		try:
			img = Image.open(path_image).convert('RGB')
		except:
			if info:
				info.info += u"ОШИБКА при загрузке изображения продукта c id=%d (проверьте правильность пути до изображения: https://%s или http://%s)\n" % (
					int(product_id), url_image, url_image)
				info.state_id = IND_STATE_ERROR
				info.save()
				return


class AdminMIndexView(TemplateView):
	"""
	Класс для отображения главной страницы
	"""
	template_name = "admin_m/index.pug"
	
	def get(self, request):
		return render(request, self.template_name, locals())


class AdminMImportInfoView(TemplateView):
	"""
	Класс для отображения страницы информации об импорте
	"""
	template_name = "admin_m/control/import/import_info.pug"
	
	def get(self, request, import_info_slug):
		current_info = get_object_or_404(ImportData, id=import_info_slug)
		current_info_information = current_info.info.replace("\n", "<br/>")
		return render(request, self.template_name, locals())


class AdminMImportView(TemplateView):
	"""
	Класс для отображения страницы импорта
	"""
	template_name = "admin_m/control/import/import.pug"
	
	def get(self, request):
		items_list = ImportData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO
		
		return render(request, self.template_name, locals())
	
	def post(self, request, template_name="admin_m/control/import/import_table.pug"):
		post_data = request.POST
		email = None
		try:
			email = post_data["email"]
		except KeyError:
			pass
		
		form = UploadFileForm(request.POST, request.FILES)
		file_name = None
		if form.is_valid():
			main_name = translit(request.FILES['file'].name.strip(), "ru", reversed=True)
			file_name = 'import/files/import_' + get_today() + '_' + main_name
			handle_uploaded_file(request.FILES['file'], file_name)
		else:
			raise Http404()
		
		try:
			type_action = post_data["type_action"]
		except KeyError:
			raise Http404()
		
		flag_hard_import = None
		try:
			flag_hard_import = post_data["flag_hard_import"]
		except KeyError:
			pass
		
		action_str = ""
		if type_action == TYPE_ACTION_IMPORT:
			if flag_hard_import:
				action_str = u"Импорт (принудительный)"
				type_action = TYPE_ACTION_IMPORT_HARD
			else:
				action_str = u"Импорт"
		elif type_action == TYPE_ACTION_CONTROL:
			action_str = u"Проверка дубликатов"
		elif type_action == TYPE_ACTION_IMPORT_MARKA:
			action_str = u"Импорт параметров"
		
		info = ImportData()
		info.name = request.FILES['file'].name
		info.action = action_str
		info.user = request.user.get_full_name()
		info.email = email
		info.file = file_name
		info.state_id = IND_STATE_WORK
		info.result = "0 %"
		info.result_percent = 0
		info.info = u""
		info.save()
		
		import threading
		t = threading.Thread(target=import_data, args=[info, request.scheme, type_action])
		t.setDaemon(True)
		t.start()
		# import_data(info, request.scheme, type_action)
		
		items_list = ImportData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO
		
		return render(request, template_name, locals())


def send_mail_result(email, user, name, link, link_admin, action, result):
	subject = u"Панель управления %s. %s (файл: %s) - %s" % (ProjectSettings.objects.first().name, action, name, result)
	plain_text = u"""Здравствуйте, %s!\n%s (файл: %s) %s.""" % (user, action, name, result)
	html_text = u"""<h1>Здравствуйте, %s!</h1>\n<p>%s файла: %s - %s. </p>\n\n
        <p>Вы можете посмотреть подробнее по <a href='%s'>прямой ссылке</a> или из <a href='%s'>панели управления</a></p>""" % (
		user, action, name, result, link, link_admin)
	
	project_settings = ProjectSettings.objects.all()
	if project_settings:
		project_settings = project_settings[0]
		from email.mime.text import MIMEText
		from email.mime.multipart import MIMEMultipart
		address = project_settings.tech_email
		# Compose message
		msg = MIMEMultipart('alternative')
		msg['From'] = project_settings.tech_email
		msg['To'] = email
		msg['Subject'] = subject
		
		part1 = MIMEText(plain_text, 'plain', 'UTF-8')
		part2 = MIMEText(html_text, 'html', 'UTF-8')
		msg.attach(part1)
		msg.attach(part2)
		
		import smtplib
		s = smtplib.SMTP(project_settings.tech_mail_server, 465)
		s.ehlo()
		s.esmtp_features["auth"] = "LOGIN PLAIN"
		s.debuglevel = 5
		s.login(project_settings.tech_email, project_settings.tech_email_pass)
		s.sendmail(address, email, msg.as_string())
		s.quit()


def get_today():
	return datetime.datetime.today().strftime("%Y-%m-%d %H:%M")


def handle_uploaded_file(f, file_name):
	filename = 'www/media/' + file_name
	with open(filename, 'wb+') as destination:
		for chunk in f.chunks():
			destination.write(chunk)


def import_data(info, scheme, type_action):
	data_load = load_xls(info)
	if info.state_id == IND_STATE_ERROR:
		info.result = DEFAULT_MESSAGE_ERROR
		info.info += DEFAULT_MESSAGE_ERROR + "\n"
		info.save()
		return
	
	data_format, new_marka_list, new_gost_list = format_type(info, data_load, type_action)
	
	if type_action == TYPE_ACTION_IMPORT:
		if info.state_id == IND_STATE_WORK:
			import_data_import(info, data_format)
	elif type_action == TYPE_ACTION_IMPORT_HARD:
		if info.state_id == IND_STATE_WORK:
			import_data_import_hard(info, data_format)
	elif type_action == TYPE_ACTION_CONTROL:
		import_data_control(info, data_format)
	elif type_action == TYPE_ACTION_IMPORT_MARKA:
		import_data_import_marka(info, new_marka_list, new_gost_list)
	
	if info.state_id == IND_STATE_ERROR:
		info.result = "Завершен с ошибками"
		info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, устраните ошибки и повторите импорт"
		info.save()
	elif info.state_id == IND_STATE_POTENTIAL_ERROR:
		info.result = "Завершен с ошибками"
		info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, если Вы действительно хотите изменить категорию у существующих позиций, установите флаг 'отключить проверку дубликатов' и повторите импорт"
		info.save()
	else:
		info.state_id = IND_STATE_READY
		info.save()
	
	if info.email:
		url_site = scheme + "://" + ProjectSettings.objects.first().site_name
		link = url_site + info.get_absolute_url()
		link_admin = url_site + "/admin_m/import/"
		send_mail_result(info.email, info.user, info.name, link, link_admin, info.action, info.state.name)
		info.info += u""
		info.save()


def load_xls(info):
	file_name = info.file.path
	data_load = []
	from openpyxl import load_workbook
	wb = load_workbook(filename=file_name, read_only=True, data_only=True)
	ws = wb.active
	n = 0
	if ws.max_row < ROW_MIN_COUNT:
		info.info += u"Неверный формат файла: количество строк меньше минимального - %d\n" % ROW_MIN_COUNT
		info.state_id = IND_STATE_ERROR
	if ws.max_column != COLUMN_COUNT:
		info.info += u"Неверный формат файла: количество столбцов в файле %d, а должно быть %d\n" % (
			ws.max_column, COLUMN_COUNT)
		info.state_id = IND_STATE_ERROR
	for row in ws.rows:
		# if n == 0:
		#     n += 1
		#     continue
		current_row = []
		for cell in row:
			current_row.append(cell.value)
		if current_row and current_row[0] == None:
			break
		data_load.append(current_row)
	return data_load


def get_marka_name(marka):
	key = marka.replace(' ', '').upper()
	if (key == '5.0'):
		key = key[0:1]
		key = '0' + key + 'КП'
	if (key == u'5ПС' or key == u'3ПС'):
		key = 'СТ' + key
	if (key == u'1ПС'):
		key = 'СТ' + key
	if (key == u'08Х17Н6М2Т'):
		key = key[0:7] + key[9:10]
	if (key == u'ШХ15В'):
		key = key[0:4]
	if (key == '20.0'):
		key = '20'
	if (key == '25.0'):
		key = '25'
	if (key == '10.0'):
		key = '10'
	if (key == '15.0'):
		key = '15'
	if (key == '45.0'):
		key = '45'
	if (key == '30.0'):
		key = '30'
	if (key == '35.0'):
		key = '35'
	if (key == '40.0'):
		key = '40'
	if (key == u'СТАЛЬ55'):
		key = key[5:7]
	if (key == '50.0'):
		key = '50'
	if (key == '60.0'):
		key = '60'
	if (key == '65.0'):
		key = '65'
	if (key == '70.0'):
		key = '70'
	if (key == '0.0' or key == '0'):
		key = key[0:1]
		key = 'СТ' + key
	return key


def get_gost_name(gost, dict_gost):
	gost_5 = ""
	if not gost == "":
		key = gost.replace(' ', '')
		try:
			gost_5 = dict_gost[key]
		except:
			try:
				gost_5 = dict_gost[u"ГОСТ" + key]
			except:
				try:
					gost_5 = dict_gost[u"ГОСТР" + key]
				except:
					try:
						gost_5 = dict_gost[u"ГОСТ" + key + '-78']
					except:
						try:
							gost_5 = dict_gost[u"ГОСТ" + key + '-87']
						except:
							try:
								gost_5 = dict_gost[u"ТУ" + key]
							except:
								return False
	return gost_5


def format_type(info, data_load, type_action):
	data_format = []
	new_marka_list = []
	new_gost_list = []
	rez = True
	
	category_list = []
	category_list_map = MenuCatalog.objects.all().only('id').values('id')
	for item in category_list_map:
		category_list.append(item['id'])
	marka_dict = {}
	marka_dict_map = Marka.objects.all().only('id', 'name').values('id', 'name')
	for item in marka_dict_map:
		marka_dict[item['name'].upper().replace(' ', '')] = item['id']
	
	gost_dict = {}
	gost_dict_map = Standart.objects.all().only('id', 'name').values('id', 'name')
	for item in gost_dict_map:
		gost_dict[item['name'].replace(' ', '')] = item['id']
	
	info.result = "Подготовка"
	# info.result = "Подготовка - 0 %"
	# count_items = len(data_load) - 1
	for i in range(1, len(data_load)):
		# info.result_percent += (1. / count_items * 100)
		# print("format %.2f  %d / %d" % (info.result_percent, i, count_items))
		# if i % COUNT_ROW_FOR_INFO_UPDATE_FORMAT == 0 or i == count_items:
		#     info.result = ("Подготовка - %.2f" % info.result_percent) + " %"
		#     info.save()
		
		data_format.append([])
		data_format[i - 1].append(data_load[i][COLUMN_ID])
		data_format[i - 1].append(data_load[i][COLUMN_ORDER])
		data_format[i - 1].append(data_load[i][COLUMN_NAME])
		
		category_id = ""
		tmp_category_id = ""
		category_id = data_load[i][COLUMN_CATEGORY_ID]
		
		try:
			tmp_category_id = int(category_id)
			category_id = tmp_category_id
		except:
			pass
		
		if not (category_id in category_list):
			info.info += u"%d строка: Отсутствует категория с идентификатором: %d\n" % (i, category_id)
			info.state_id = IND_STATE_ERROR
			info.save()
		
		data_format[i - 1].append(category_id)
		data_format[i - 1].append(data_load[i][COLUMN_SIZE_A])
		data_format[i - 1].append(data_load[i][COLUMN_SIZE_B])
		data_format[i - 1].append(data_load[i][COLUMN_SIZE_C])
		data_format[i - 1].append(data_load[i][COLUMN_SIZE_D])
		data_format[i - 1].append(data_load[i][COLUMN_SIZE_E])
		data_format[i - 1].append(data_load[i][COLUMN_SIZE_F])
		data_format[i - 1].append(data_load[i][COLUMN_SIZE_L])
		
		data_load[i][COLUMN_MARKA] = str(data_load[i][COLUMN_MARKA])
		marka_name = get_marka_name(data_load[i][COLUMN_MARKA])
		if marka_name:
			marka_id = ''
			try:
				marka_id = marka_dict[marka_name]
				data_format[i - 1].append(marka_id)
			except:
				# try:
				info.info += u"%d строка: Отсутствует марка: %s\n" % (i, data_load[i][COLUMN_MARKA])
				# except:
				#     info.info += str(i) + " строка: Отсутствует марка содержащая спецсимволы\n"
				info.state_id = IND_STATE_ERROR
				# info.save()
				if not (data_load[i][COLUMN_MARKA] in new_marka_list):
					new_marka_list.append(data_load[i][COLUMN_MARKA])
				data_format[i - 1].append('')
		else:
			data_format[i - 1].append(marka_name)
		
		if data_load[i][COLUMN_GOST]:
			gost_id = get_gost_name(data_load[i][COLUMN_GOST], gost_dict)
			if gost_id:
				data_format[i - 1].append(gost_id)
			else:
				info.info += u"%d строка: Отсутствует ГОСТ, ТУ: %s\n" % (i, data_load[i][COLUMN_GOST])
				info.state_id = IND_STATE_ERROR
				info.save()
				if not (data_load[i][COLUMN_GOST] in new_gost_list):
					new_gost_list.append(data_load[i][COLUMN_GOST])
		else:
			data_format[i - 1].append('')
		
		data_format[i - 1].append(data_load[i][COLUMN_AVAILABLE])
		data_format[i - 1].append(data_load[i][COLUMN_PRICE])
		
		# if not data_load[i][COLUMN_PRICE]:
		#     info.info += u"%d строка: Отсутствует цена!\n" % (i)
		#     info.state_id = IND_STATE_ERROR
		#     info.save()
		
		data_format[i - 1].append(data_load[i][COLUMN_ED_IZM])
		
		image_name_1 = ""
		image_name_1_http = data_load[i][COLUMN_IMAGE_1]
		image_name_2 = ""
		image_name_2_http = data_load[i][COLUMN_IMAGE_2]
		
		if data_load[i][COLUMN_IMAGE_1]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_1])
			if rezImg[-1]:
				image_name_1 = rezImg[-1]
			elif rezImg[-2]:
				image_name_1 = rezImg[-2]
		
		if data_load[i][COLUMN_IMAGE_2]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_2])
			if rezImg[-1]:
				image_name_2 = rezImg[-1]
			elif rezImg[-2]:
				image_name_2 = rezImg[-2]
		
		image_name_bd = "uploads/images/load/" + image_name_1
		data_format[i - 1].append(image_name_bd)
		image_name_bd = "uploads/images/load/" + image_name_2
		data_format[i - 1].append(image_name_bd)
		
		data_format[i - 1].append(data_load[i][COLUMN_DESCRIPTION])
		data_format[i - 1].append(data_load[i][COLUMN_VENDOR])
		data_format[i - 1].append(data_load[i][COLUMN_COUNTRY_VENDOR])
		data_format[i - 1].append(data_load[i][COLUMN_METHOD])
		
		data_format[i - 1].append(data_load[i][COLUMN_SERTIFICATE_ID])
		data_format[i - 1].append(data_load[i][COLUMN_PRODUCT_1_ID])
		data_format[i - 1].append(data_load[i][COLUMN_PRODUCT_2_ID])
		data_format[i - 1].append(data_load[i][COLUMN_PRODUCT_3_ID])
		data_format[i - 1].append(data_load[i][COLUMN_PRODUCT_4_ID])
		data_format[i - 1].append(data_load[i][COLUMN_USLUGA_1_ID])
		data_format[i - 1].append(data_load[i][COLUMN_USLUGA_2_ID])
		data_format[i - 1].append(data_load[i][COLUMN_USLUGA_3_ID])
		data_format[i - 1].append(data_load[i][COLUMN_USLUGA_4_ID])
		
		data_format[i - 1].append(image_name_1)
		data_format[i - 1].append(image_name_2)
		data_format[i - 1].append(image_name_1_http)
		data_format[i - 1].append(image_name_2_http)
	
	info.result_percent = 0
	info.save()
	if new_marka_list:
		info.info += u"\nОтсутствуют марки:\n"
		for item in new_marka_list:
			info.info += u"%s\n" % item
		info.save()
	if new_gost_list:
		info.info += u"\nОтсутствуют ГОСТы:\n"
		for item in new_gost_list:
			info.info += u"%s\n" % item
		info.save()
	info.info += u"Вы можете загрузить их через 'Импорт параметров'\n"
	info.save()
	return data_format, new_marka_list, new_gost_list


def import_data_import_hard(info, data_format):
	list_import_image = []
	count_products = len(data_format)
	for i in range(len(data_format)):
		product = Product()
		info.result_percent += (1. / count_products * 100)
		# print("import hard %.2f   %d/%d" % (info.result_percent, i, count_products))
		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i + 1 == count_products:
			info.result = ("%.2f" % info.result_percent) + " %"
			info.save()
		
		product.id = data_format[i][COLUMN_ID]
		if data_format[i][1]:
			product.order_number = data_format[i][1]
		
		product.name = data_format[i][2]
		
		product.catalog_id = data_format[i][3]
		if data_format[i][4]:
			product.param_1 = str(data_format[i][4])
		if data_format[i][5]:
			product.param_2 = str(data_format[i][5])
		if data_format[i][6]:
			product.param_3 = str(data_format[i][6])
		if data_format[i][7]:
			product.param_4 = str(data_format[i][7])
		if data_format[i][8]:
			product.param_5 = str(data_format[i][8])
		if data_format[i][9]:
			product.param_6 = str(data_format[i][9])
		if data_format[i][10]:
			product.param_7 = str(data_format[i][10])
		if data_format[i][11]:
			product.marka_id = data_format[i][11]
		if data_format[i][12]:
			product.standart_id = data_format[i][12]
		product.available = data_format[i][13]
		product.price = data_format[i][14]
		product.ed_izm = data_format[i][15]
		product.image = data_format[i][16]
		product.image_dop = data_format[i][17]
		product.description = data_format[i][18]
		product.vendor = data_format[i][19]
		product.vendor_country = data_format[i][20]
		product.method_of_manufacture = data_format[i][21]
		product.isHidden = False
		if data_format[i][22]:
			product.sertificate_id = data_format[i][22]
		
		# if product.slug is None or product.slug == "":
		product.set_slug()
		
		# Контроль на дубли позиций по латинским именам
		try:
			product_tmp = Product.objects.only('id').get(slug=product.slug)
			if product_tmp.id != int(product.id):
				info.info += u"ОШИБКА при сохранении продукта - дубликат латинского имени '%s' (текущий id=%d, существующий id=%d)\n" % (
					product.slug, int(product.id), product_tmp.id)
				info.state_id = IND_STATE_ERROR
				info.save()
				break
		except:
			pass
		
		try:
			product.save()
		except Exception as e:
			info.info += u"ОШИБКА при сохранении продукта id=%d: %s\n" % (int(product.id), e)
			info.state_id = IND_STATE_ERROR
			info.save()
			break
		
		# for tmp_i in range(23, 27):
		#     if data_format[i][tmp_i]:
		#         product_tableProductDop = tableProductDop()
		#         product_tableProductDop.product = product.id
		#         product_tableProductDop.product_dop = data_format[i][tmp_i]
		#         product_tableProductDop.save()
		# for tmp_i in range(27, 31):
		#     if data_format[i][tmp_i]:
		#         product_ProductUsluga = ProductUsluga()
		#         product_ProductUsluga.product = product.id
		#         product_ProductUsluga.usluga = data_format[i][tmp_i]
		#         product_ProductUsluga.save()
		
		if data_format[i][COLUMN_IMAGE_1_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_1_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				# print("import_image %s %s %s" % (data_format[i][COLUMN_IMAGE_1_TECH_NAME], CONTROL_CODE, data_format[i][COLUMN_IMAGE_1_HTTP]))
				import_image(None, data_format[i][COLUMN_IMAGE_1_TECH_NAME], data_format[i][COLUMN_IMAGE_1_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)
		
		if data_format[i][COLUMN_IMAGE_2_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_2_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				# print("import_image %s %s %s" % (data_format[i][COLUMN_IMAGE_2_TECH_NAME], CONTROL_CODE, data_format[i][COLUMN_IMAGE_2_HTTP]))
				import_image(None, data_format[i][COLUMN_IMAGE_2_TECH_NAME], data_format[i][COLUMN_IMAGE_2_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)
		
		if info.state_id == IND_STATE_ERROR:
			break
	
	if info.state_id != IND_STATE_ERROR:
		info.result = "100 %"
		info.save()


def import_data_import(info, data_format):
	info.result = "Проверка id категорий - 0 %"
	count_products = len(data_format)
	for i in range(len(data_format)):
		info.result_percent += (1. / count_products * 100)
		# print("control id %.2f   %d/%d" % (info.result_percent, i, count_products))
		if i % COUNT_ROW_FOR_INFO_UPDATE_CONTROL == 0 or i + 1 == count_products:
			info.result = ("Проверка id категорий - %.2f" % info.result_percent) + " %"
			info.save()
		
		try:
			product = Product.objects.only('id').get(id=data_format[i][COLUMN_ID])
		except:
			continue
		
		if product.catalog_id != data_format[i][3]:
			info.info += u"%d строка: Внимание! Возможно ошибочно указан id категории %d (сейчас на сайте у данного продукта id категории %d) у позиции id=%d\n" % (
				i, data_format[i][3], product.catalog_id, data_format[i][COLUMN_ID])
			if info.state_id == IND_STATE_WORK:
				info.state_id = IND_STATE_POTENTIAL_ERROR
	info.result_percent = 0
	info.save()
	if info.state_id == IND_STATE_WORK:
		import_data_import_hard(info, data_format)


def import_data_control(info, data_format):
	count_products = len(data_format)
	for i in range(len(data_format)):
		info.result_percent += (1. / count_products * 100)
		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i + 1 == count_products:
			info.result = ("%.2f" % info.result_percent) + " %"
			info.save()
		
		try:
			product = Product.objects.get(id=data_format[i][COLUMN_ID])
		except:
			continue
		
		info.info += u"%d строка: Указанный id уже существует (id - %s) \n" % (i, data_format[i][COLUMN_ID])
		info.state_id = IND_STATE_ERROR
		info.save()


def import_data_import_marka(info, new_marka_list, new_gost_list):
	for item in new_marka_list:
		marka = Marka()
		marka.name = item
		marka.set_slug()
		marka.save()
	for item in new_gost_list:
		marka = Standart()
		marka.name = item
		marka.set_slug()
		marka.save()
	
	info.result = "100 %"
	info.state_id = IND_STATE_WORK
	info.info += u"\nМарки успешно загружены"
	info.save()
